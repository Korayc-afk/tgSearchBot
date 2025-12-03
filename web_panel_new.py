"""
Telegram Monitoring Bot - Web Panel (Multi-Tenant)
Flask tabanlı web arayüzü - Çoklu grup desteği
"""

from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, url_for, session
from flask_cors import CORS
from flask_login import login_user, logout_user, login_required, current_user
import asyncio
import json
import os
import subprocess
import threading
import logging
import traceback
from datetime import datetime, timedelta
from telethon import TelegramClient
from database import init_db, create_super_admin, SessionLocal, User, Tenant, TenantConfig, Result, MessageStatistics, UserTenant
from auth import login_manager, verify_password, require_super_admin, require_tenant_access
from tenant_manager import (
    create_tenant, get_tenant, get_tenant_by_slug, get_user_tenants,
    update_tenant, delete_tenant, get_tenant_config, update_tenant_config,
    add_user_to_tenant, remove_user_from_tenant, get_tenant_users
)

# Logging yapılandırması
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='.')
app.secret_key = os.environ.get('SECRET_KEY', 'padisah-telegram-monitoring-secret-key-change-in-production')
CORS(app)

# Request logging middleware
@app.before_request
def log_request_info():
    """Her request'i logla"""
    try:
        logger.info(f"🔵 REQUEST: {request.method} {request.path}")
        logger.info(f"   Headers: {dict(request.headers)}")
        if request.is_json:
            try:
                logger.info(f"   JSON Body: {json.dumps(request.json, indent=2, ensure_ascii=False)}")
            except:
                logger.info(f"   JSON Body: (parse edilemedi)")
        elif request.form:
            logger.info(f"   Form Data: {dict(request.form)}")
        elif request.args:
            logger.info(f"   Query Params: {dict(request.args)}")
        try:
            if current_user.is_authenticated:
                logger.info(f"   User: {current_user.username} (ID: {current_user.id}, Role: {current_user.role})")
        except:
            logger.info(f"   User: (yüklenemedi)")
    except Exception as e:
        logger.error(f"   Logging hatası: {e}")

@app.after_request
def log_response_info(response):
    """Her response'u logla"""
    logger.info(f"🟢 RESPONSE: {request.method} {request.path} - Status: {response.status_code}")
    return response

# Error handlers
@app.errorhandler(400)
def bad_request(error):
    """400 Bad Request handler"""
    logger.error(f"❌ BAD REQUEST: {request.method} {request.path}")
    logger.error(f"   Error: {str(error)}")
    logger.error(f"   Error Type: {type(error).__name__}")
    logger.error(f"   Request Data: {request.get_data(as_text=True)}")
    logger.error(f"   Request Args: {dict(request.args)}")
    logger.error(f"   Request Form: {dict(request.form)}")
    try:
        if request.is_json:
            logger.error(f"   Request JSON: {json.dumps(request.json, indent=2, ensure_ascii=False)}")
    except:
        logger.error(f"   Request JSON: (parse edilemedi)")
    logger.error(f"   Traceback: {traceback.format_exc()}")
    
    # Daha açıklayıcı hata mesajı
    error_msg = str(error)
    if 'tenant_id' in error_msg.lower() or 'tenant' in request.path.lower():
        error_msg = "Tenant ID bulunamadı veya geçersiz!"
    elif 'json' in error_msg.lower():
        error_msg = "JSON formatı geçersiz!"
    elif 'form' in error_msg.lower():
        error_msg = "Form verisi eksik veya geçersiz!"
    
    return jsonify({
        'success': False,
        'message': f'Bad Request: {error_msg}',
        'details': {
            'method': request.method,
            'path': request.path,
            'error': str(error),
            'error_type': type(error).__name__
        }
    }), 400

@app.errorhandler(404)
def not_found(error):
    """404 Not Found handler"""
    logger.warning(f"⚠️  NOT FOUND: {request.method} {request.path}")
    return jsonify({
        'success': False,
        'message': f'Endpoint bulunamadı: {request.path}'
    }), 404

@app.errorhandler(500)
def internal_error(error):
    """500 Internal Server Error handler"""
    logger.error(f"❌ INTERNAL ERROR: {request.method} {request.path}")
    logger.error(f"   Error: {str(error)}")
    logger.error(f"   Traceback: {traceback.format_exc()}")
    
    return jsonify({
        'success': False,
        'message': f'Internal Server Error: {str(error)}',
        'details': traceback.format_exc()
    }), 500

@app.errorhandler(Exception)
def handle_exception(e):
    """Genel exception handler"""
    logger.error(f"❌ EXCEPTION: {request.method} {request.path}")
    logger.error(f"   Exception Type: {type(e).__name__}")
    logger.error(f"   Exception Message: {str(e)}")
    logger.error(f"   Traceback: {traceback.format_exc()}")
    
    return jsonify({
        'success': False,
        'message': f'Hata: {str(e)}',
        'error_type': type(e).__name__,
        'details': traceback.format_exc()
    }), 500

# Flask-Login'i başlat
login_manager.init_app(app)

# Bot process tracking (tenant bazlı)
bot_processes = {}  # {tenant_id: process}
bot_statuses = {}  # {tenant_id: status}
bot_logs = {}  # {tenant_id: [logs]}

# ==================== HELPER FUNCTIONS ====================

def get_current_tenant_id():
    """Mevcut kullanıcının tenant ID'sini al"""
    # Önce request'ten al (args veya json)
    tenant_id = None
    try:
        tenant_id = request.args.get('tenant_id')
        if tenant_id:
            try:
                return int(tenant_id)
            except (ValueError, TypeError):
                pass
    except:
        pass
    
    try:
        if request.is_json and request.json:
            tenant_id = request.json.get('tenant_id')
            if tenant_id:
                try:
                    return int(tenant_id)
                except (ValueError, TypeError):
                    pass
    except:
        pass
    
    # Request'te yoksa session'dan al
    if not tenant_id:
        try:
            tenant_id = session.get('selected_tenant_id')
            if tenant_id:
                return int(tenant_id)
        except:
            pass
    
    # Session'da da yoksa kullanıcının tenant'larından al
    if not tenant_id:
        if current_user.is_super_admin:
            # Süper admin ise ilk aktif tenant'ı al
            db = SessionLocal()
            try:
                first_tenant = db.query(Tenant).filter_by(is_active=True).first()
                if first_tenant:
                    tenant_id = first_tenant.id
                    session['selected_tenant_id'] = tenant_id
            finally:
                db.close()
        else:
            # Normal kullanıcı için ilk tenant'ını al
            user_tenants = get_user_tenants(current_user.id)
            if user_tenants:
                tenant_id = user_tenants[0].id
                session['selected_tenant_id'] = tenant_id
    
    return tenant_id

def get_telegram_client_for_tenant(tenant_id):
    """Tenant için Telegram client oluştur"""
    config = get_tenant_config(tenant_id)
    if not config or not config.api_id or not config.get_api_hash():
        return None
    
    # Tenant slug'ını al (session içinde)
    db = SessionLocal()
    try:
        tenant = db.query(Tenant).filter_by(id=tenant_id).first()
        if not tenant:
            return None
        tenant_slug = tenant.slug
    finally:
        db.close()
    
    session_path = config.session_file_path or f'tenants/{tenant_slug}/session.session'
    
    # Session dosyası için mutlak yol kullan
    if not os.path.isabs(session_path):
        session_path = os.path.abspath(session_path)
    
    session_name = session_path.replace('.session', '')
    
    # Session dosyasının dizinini kontrol et ve oluştur
    session_dir = os.path.dirname(session_name)
    if session_dir and not os.path.exists(session_dir):
        os.makedirs(session_dir, exist_ok=True)
    
    return TelegramClient(session_name, config.api_id, config.get_api_hash())

# ==================== AUTHENTICATION ROUTES ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login sayfası - Normal kullanıcılar için (username, password, tenant seçimi)"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        tenant_id = request.form.get('tenant_id')
        
        # Süper admin kontrolü - süper admin ise tenant seçimi olmadan giriş yapabilir
        user = verify_password(username, password)
        if user:
            if user.is_super_admin:
                # Süper admin direkt super-admin paneline gitsin
                login_user(user, remember=True)
                return redirect(url_for('super_admin_dashboard'))
            else:
                # Normal kullanıcı için tenant_id gerekli
                if not tenant_id:
                    db = SessionLocal()
                    try:
                        tenants = db.query(Tenant).filter_by(is_active=True).all()
                        tenant_list = [{'id': t.id, 'name': t.name} for t in tenants]
                    finally:
                        db.close()
                    return render_template('login.html', error='Lütfen bir grup seçin!', tenants=tenant_list)
                
                # Kullanıcının bu tenant'a erişimi var mı?
                db = SessionLocal()
                try:
                    user_tenant = db.query(UserTenant).filter_by(
                        user_id=user.id,
                        tenant_id=int(tenant_id)
                    ).first()
                    if not user_tenant:
                        tenants = db.query(Tenant).filter_by(is_active=True).all()
                        tenant_list = [{'id': t.id, 'name': t.name} for t in tenants]
                        return render_template('login.html', error='Bu gruba erişim yetkiniz yok!', tenants=tenant_list)
                finally:
                    db.close()
                
                # Giriş başarılı - tenant_id'yi session'a kaydet
                login_user(user, remember=True)
                session['selected_tenant_id'] = int(tenant_id)
                return redirect(url_for('index'))
        else:
            # Hatalı giriş - tenant listesini tekrar göster
            db = SessionLocal()
            try:
                tenants = db.query(Tenant).filter_by(is_active=True).all()
                tenant_list = [{'id': t.id, 'name': t.name} for t in tenants]
            finally:
                db.close()
            return render_template('login.html', error='Kullanıcı adı veya şifre hatalı!', tenants=tenant_list)
    
    if current_user.is_authenticated:
        if current_user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        return redirect(url_for('index'))
    
    # GET request - tenant listesini al
    db = SessionLocal()
    try:
        tenants = db.query(Tenant).filter_by(is_active=True).all()
        tenant_list = [{'id': t.id, 'name': t.name} for t in tenants]
    finally:
        db.close()
    
    return render_template('login.html', tenants=tenant_list)

@app.route('/logout')
@login_required
def logout():
    """Çıkış yap"""
    logout_user()
    return redirect(url_for('login'))

@app.route('/logoSeffaf.png')
def serve_logo_seffaf():
    """Şeffaf logo dosyasını serve et"""
    return send_from_directory('.', 'logoSeffaf.png')

@app.route('/s-l400.jpg')
def serve_logo():
    """Logo dosyasını serve et"""
    return send_from_directory('.', 's-l400.jpg')

# ==================== MAIN ROUTES ====================

@app.route('/')
@login_required
def index():
    """Ana sayfa - Herkes aynı paneli görür, sadece veriler farklı"""
    # Kullanıcının tenant'ını belirle
    tenant_id = None
    tenant_name = None
    
    if current_user.is_super_admin:
        # Süper admin ise tüm tenant'ları görebilir, varsayılan olarak ilkini seç
        db = SessionLocal()
        try:
            first_tenant = db.query(Tenant).filter_by(is_active=True).first()
            if first_tenant:
                tenant_id = first_tenant.id
                tenant_name = first_tenant.name
                # Expunge yap
                db.expunge(first_tenant)
        finally:
            db.close()
    else:
        # Normal admin için ilk tenant'ını al
        user_tenants = get_user_tenants(current_user.id)
        if user_tenants:
            tenant_id = user_tenants[0].id
            tenant_name = user_tenants[0].name
        else:
            return render_template('no_tenant.html')
    
    return render_template('index.html', tenant_id=tenant_id, tenant_name=tenant_name or 'Telegram Monitoring', is_super_admin=current_user.is_super_admin)

@app.route('/super-admin')
@login_required
@require_super_admin
def super_admin_dashboard():
    """Süper admin dashboard"""
    return render_template('super_admin.html')

@app.route('/admin/<int:tenant_id>')
@login_required
@require_tenant_access('tenant_id')
def admin_dashboard(tenant_id):
    """Normal admin dashboard"""
    tenant = get_tenant(tenant_id)
    if not tenant:
        return redirect(url_for('index'))
    return render_template('admin.html', tenant_id=tenant_id, tenant_name=tenant.name)

# ==================== SUPER ADMIN API ROUTES ====================

@app.route('/api/super-admin/dashboard')
@login_required
@require_super_admin
def super_admin_dashboard_data():
    """Süper admin dashboard verileri"""
    db = SessionLocal()
    try:
        # Tüm tenant'lar
        tenants = db.query(Tenant).filter_by(is_active=True).all()
        
        # İstatistikler
        total_tenants = len(tenants)
        total_users = db.query(User).count()
        total_results = db.query(Result).count()
        
        # Son 7 günün istatistikleri
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_results = db.query(Result).filter(Result.timestamp >= seven_days_ago).count()
        
        # Tenant bazında istatistikler
        tenant_stats = []
        for tenant in tenants:
            # Tenant bilgilerini session içinde al
            tenant_id = tenant.id
            tenant_name = tenant.name
            tenant_slug = tenant.slug
            tenant_created_at = tenant.created_at
            
            tenant_result_count = db.query(Result).filter_by(tenant_id=tenant_id).count()
            tenant_stats.append({
                'id': tenant_id,
                'name': tenant_name,
                'slug': tenant_slug,
                'result_count': tenant_result_count,
                'created_at': tenant_created_at.isoformat() if tenant_created_at else None
            })
        
        return jsonify({
            'success': True,
            'stats': {
                'total_tenants': total_tenants,
                'total_users': total_users,
                'total_results': total_results,
                'recent_results': recent_results
            },
            'tenants': [{
                'id': t.id,
                'name': t.name,
                'slug': t.slug,
                'is_active': t.is_active,
                'created_at': t.created_at.isoformat() if t.created_at else None
            } for t in tenants],
            'tenant_stats': tenant_stats
        })
    finally:
        db.close()

@app.route('/api/super-admin/tenants', methods=['GET'])
@login_required
@require_super_admin
def list_tenants():
    """Tüm tenant'ları listele"""
    db = SessionLocal()
    try:
        tenants = db.query(Tenant).all()
        tenant_list = []
        for t in tenants:
            # Her tenant için sonuç sayısını hesapla
            result_count = db.query(Result).filter_by(tenant_id=t.id).count()
            tenant_list.append({
                'id': t.id,
                'name': t.name,
                'slug': t.slug,
                'is_active': t.is_active,
                'created_at': t.created_at.isoformat() if t.created_at else None,
                'created_by': t.created_by,
                'result_count': result_count
            })
        return jsonify({
            'success': True,
            'tenants': tenant_list
        })
    finally:
        db.close()

@app.route('/api/super-admin/tenants', methods=['POST'])
@login_required
@require_super_admin
def create_tenant_api():
    """Yeni tenant oluştur"""
    try:
        logger.info("📥 POST /api/super-admin/tenants çağrıldı")
        data = request.json
        logger.info(f"   Request Data: {json.dumps(data, indent=2, ensure_ascii=False)}")
        
        name = data.get('name', '').strip()
        
        if not name:
            logger.warning("   ⚠️  Grup adı eksik!")
            return jsonify({'success': False, 'message': 'Grup adı gerekli!'})
        
        # Tenant oluştur (sadece tenant, user_tenant ilişkisi oluşturma)
        tenant = create_tenant(name, current_user.id)
        # Expunge edilmiş tenant'tan bilgileri al (expunge etmeden önce alındığı için güvenli)
        try:
            tenant_id = tenant.id
            tenant_name = tenant.name
            tenant_slug = tenant.slug
            logger.info(f"   ✅ Tenant oluşturuldu: {tenant_id} - {tenant_name}")
        except Exception as e:
            # Eğer expunge edilmiş tenant'tan bilgi alınamazsa, database'den tekrar al
            logger.warning(f"   ⚠️  Tenant bilgileri alınamadı, database'den tekrar alınıyor: {e}")
            db = SessionLocal()
            try:
                tenant_db = db.query(Tenant).filter_by(name=name).order_by(Tenant.id.desc()).first()
                if tenant_db:
                    tenant_id = tenant_db.id
                    tenant_name = tenant_db.name
                    tenant_slug = tenant_db.slug
                else:
                    raise Exception("Tenant database'de bulunamadı!")
            finally:
                db.close()
        
        return jsonify({
            'success': True,
            'message': 'Grup başarıyla oluşturuldu!',
            'tenant': {
                'id': tenant_id,
                'name': tenant_name,
                'slug': tenant_slug
            }
        })
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/tenants/<int:tenant_id>', methods=['PUT'])
@login_required
@require_super_admin
def update_tenant_api(tenant_id):
    """Tenant'ı güncelle"""
    try:
        data = request.json
        name = data.get('name')
        is_active = data.get('is_active')
        
        tenant = update_tenant(tenant_id, name=name, is_active=is_active)
        if tenant:
            return jsonify({'success': True, 'message': 'Grup güncellendi!'})
        else:
            return jsonify({'success': False, 'message': 'Grup bulunamadı!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/tenants/<int:tenant_id>', methods=['DELETE'])
@login_required
@require_super_admin
def delete_tenant_api(tenant_id):
    """Tenant'ı sil"""
    try:
        if delete_tenant(tenant_id):
            return jsonify({'success': True, 'message': 'Grup silindi!'})
        else:
            return jsonify({'success': False, 'message': 'Grup bulunamadı!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/users', methods=['GET'])
@login_required
@require_super_admin
def list_users():
    """Tüm kullanıcıları listele (şifreler dahil)"""
    db = SessionLocal()
    try:
        users = db.query(User).all()
        return jsonify({
            'success': True,
            'users': [{
                'id': u.id,
                'username': u.username,
                'password': u.password_plain or '***',  # Şifreyi göster (plain text)
                'role': u.role,
                'last_login': u.last_login.isoformat() if u.last_login else None,
                'created_at': u.created_at.isoformat() if u.created_at else None
            } for u in users]
        })
    finally:
        db.close()

@app.route('/api/super-admin/users/<int:user_id>/tenants', methods=['GET'])
@login_required
@require_super_admin
def get_user_tenants_api(user_id):
    """Kullanıcının grup bilgilerini al"""
    db = SessionLocal()
    try:
        user_tenants = db.query(UserTenant).filter_by(user_id=user_id).all()
        tenant_ids = [ut.tenant_id for ut in user_tenants]
        return jsonify({
            'success': True,
            'tenants': tenant_ids
        })
    finally:
        db.close()

@app.route('/api/super-admin/users/<int:user_id>/tenants', methods=['PUT'])
@login_required
@require_super_admin
def update_user_tenants_api(user_id):
    """Kullanıcının grup ilişkilerini güncelle"""
    try:
        data = request.json
        tenant_ids = data.get('tenant_ids', [])
        
        db = SessionLocal()
        try:
            # Mevcut ilişkileri sil
            db.query(UserTenant).filter_by(user_id=user_id).delete()
            
            # Yeni ilişkileri ekle
            for tenant_id in tenant_ids:
                user_tenant = UserTenant(
                    user_id=user_id,
                    tenant_id=tenant_id,
                    role='owner'
                )
                db.add(user_tenant)
            
            db.commit()
            return jsonify({'success': True, 'message': 'Grup ilişkileri güncellendi!'})
        except Exception as e:
            db.rollback()
            return jsonify({'success': False, 'message': f'Hata: {str(e)}'})
        finally:
            db.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/users', methods=['POST'])
@login_required
@require_super_admin
def create_user():
    """Yeni kullanıcı oluştur"""
    try:
        from hashlib import sha256
        
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        role = data.get('role', 'admin')
        tenant_ids = data.get('tenant_ids', [])  # Kullanıcının erişebileceği tenant'lar
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Kullanıcı adı ve şifre gerekli!'})
        
        db = SessionLocal()
        try:
            # Kullanıcı zaten var mı?
            existing = db.query(User).filter_by(username=username).first()
            if existing:
                return jsonify({'success': False, 'message': 'Bu kullanıcı adı zaten kullanılıyor!'})
            
            # Yeni kullanıcı oluştur
            password_hash = sha256(password.encode()).hexdigest()
            user = User(
                username=username, 
                password_hash=password_hash, 
                password_plain=password,  # Şifreyi plain text olarak sakla (sadece super admin görebilir)
                role=role
            )
            db.add(user)
            db.flush()  # ID'yi almak için
            
            # Tenant'lara ekle (aynı session içinde, user commit edilmeden önce)
            for tenant_id in tenant_ids:
                user_tenant = UserTenant(
                    user_id=user.id,
                    tenant_id=tenant_id,
                    role='owner'
                )
                db.add(user_tenant)
            
            db.commit()
            return jsonify({'success': True, 'message': 'Kullanıcı oluşturuldu!'})
        except Exception as e:
            db.rollback()
            return jsonify({'success': False, 'message': f'Hata: {str(e)}'})
        finally:
            db.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/users/<int:user_id>', methods=['PUT'])
@login_required
@require_super_admin
def update_user(user_id):
    """Kullanıcıyı güncelle"""
    try:
        from hashlib import sha256
        
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        role = data.get('role')
        
        db = SessionLocal()
        try:
            user = db.query(User).filter_by(id=user_id).first()
            if not user:
                return jsonify({'success': False, 'message': 'Kullanıcı bulunamadı!'})
            
            # Kullanıcı adı güncelle
            if username and username != user.username:
                # Yeni kullanıcı adı zaten kullanılıyor mu?
                existing = db.query(User).filter_by(username=username).first()
                if existing and existing.id != user_id:
                    return jsonify({'success': False, 'message': 'Bu kullanıcı adı zaten kullanılıyor!'})
                user.username = username
            
            # Şifre güncelle
            if password:
                password_hash = sha256(password.encode()).hexdigest()
                user.password_hash = password_hash
                user.password_plain = password  # Plain text olarak da sakla
            
            # Rol güncelle
            if role:
                user.role = role
            
            db.commit()
            return jsonify({'success': True, 'message': 'Kullanıcı güncellendi!'})
        except Exception as e:
            db.rollback()
            return jsonify({'success': False, 'message': f'Hata: {str(e)}'})
        finally:
            db.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/users/<int:user_id>', methods=['DELETE'])
@login_required
@require_super_admin
def delete_user(user_id):
    """Kullanıcıyı sil"""
    try:
        db = SessionLocal()
        try:
            user = db.query(User).filter_by(id=user_id).first()
            if not user:
                return jsonify({'success': False, 'message': 'Kullanıcı bulunamadı!'})
            
            # Süper admin kendini silemez
            if user.role == 'super_admin' and user.id == current_user.id:
                return jsonify({'success': False, 'message': 'Kendi hesabınızı silemezsiniz!'})
            
            db.delete(user)
            db.commit()
            return jsonify({'success': True, 'message': 'Kullanıcı silindi!'})
        except Exception as e:
            db.rollback()
            return jsonify({'success': False, 'message': f'Hata: {str(e)}'})
        finally:
            db.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/super-admin/tenants/<int:tenant_id>/results')
@login_required
@require_super_admin
def get_tenant_results(tenant_id):
    """Tenant'ın sonuçlarını al (süper admin)"""
    db = SessionLocal()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = db.query(Result).filter_by(tenant_id=tenant_id)
        
        if start_date:
            query = query.filter(Result.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Result.timestamp <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
        
        results = query.order_by(Result.timestamp.desc()).limit(1000).all()
        
        return jsonify({
            'success': True,
            'results': [{
                'id': r.id,
                'timestamp': r.timestamp.isoformat(),
                'group_name': r.group_name,
                'group_id': r.group_id,
                'message_text': r.message_text,
                'found_keywords': r.found_keywords,
                'found_links': r.found_links,
                'message_link': r.message_link,
                'views_count': r.views_count,
                'forwards_count': r.forwards_count,
                'reactions_count': r.reactions_count,
                'reactions_detail': r.reactions_detail,
                'replies_count': r.replies_count
            } for r in results]
        })
    finally:
        db.close()

# ==================== ADMIN API ROUTES ====================

@app.route('/api/admin/<int:tenant_id>/config', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def get_tenant_config_api(tenant_id):
    """Tenant config'ini al"""
    config = get_tenant_config(tenant_id)
    if not config:
        return jsonify({'success': False, 'message': 'Config bulunamadı!'})
    
    return jsonify({
        'success': True,
        'config': {
            'api_id': config.api_id,
            'api_hash': '***' if config.api_hash_encrypted else '',
            'phone_number': config.phone_number,
            'group_ids': config.group_ids or [],
            'search_keywords': config.search_keywords or [],
            'search_links': config.search_links or [],
            'scan_time_range': config.scan_time_range or '7days'
        }
    })

@app.route('/api/admin/<int:tenant_id>/config', methods=['POST'])
@login_required
@require_tenant_access('tenant_id')
def save_tenant_config_api(tenant_id):
    """Tenant config'ini kaydet"""
    try:
        data = request.json
        
        update_data = {}
        if 'api_id' in data:
            update_data['api_id'] = data['api_id']
        if 'api_hash' in data and data['api_hash'] != '***':
            update_data['api_hash'] = data['api_hash']
        if 'phone_number' in data:
            update_data['phone_number'] = data['phone_number']
        if 'group_ids' in data:
            update_data['group_ids'] = data['group_ids']
        if 'search_keywords' in data:
            update_data['search_keywords'] = [kw.strip() for kw in data['search_keywords'] if kw.strip()]
        if 'search_links' in data:
            update_data['search_links'] = [link.strip() for link in data['search_links'] if link.strip()]
        if 'scan_time_range' in data:
            update_data['scan_time_range'] = data['scan_time_range']
        
        config = update_tenant_config(tenant_id, **update_data)
        if config:
            return jsonify({'success': True, 'message': 'Ayarlar kaydedildi!'})
        else:
            return jsonify({'success': False, 'message': 'Config bulunamadı!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/admin/<int:tenant_id>/results', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def get_results_api(tenant_id):
    """Sonuçları al"""
    db = SessionLocal()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = int(request.args.get('limit', 100))
        
        query = db.query(Result).filter_by(tenant_id=tenant_id)
        
        if start_date:
            query = query.filter(Result.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Result.timestamp <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
        
        results = query.order_by(Result.timestamp.desc()).limit(limit).all()
        
        return jsonify({
            'success': True,
            'results': [{
                'id': r.id,
                'timestamp': r.timestamp.isoformat(),
                'group_name': r.group_name,
                'group_id': r.group_id,
                'message_text': r.message_text,
                'found_keywords': r.found_keywords,
                'found_links': r.found_links,
                'message_link': r.message_link,
                'views_count': r.views_count,
                'forwards_count': r.forwards_count,
                'reactions_count': r.reactions_count,
                'reactions_detail': r.reactions_detail,
                'replies_count': r.replies_count
            } for r in results]
        })
    finally:
        db.close()

@app.route('/api/admin/<int:tenant_id>/statistics', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def get_statistics_api(tenant_id):
    """İstatistikleri al"""
    db = SessionLocal()
    try:
        days = int(request.args.get('days', 30))
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Günlük istatistikler
        daily_stats = db.query(MessageStatistics).filter(
            MessageStatistics.tenant_id == tenant_id,
            MessageStatistics.date >= start_date
        ).order_by(MessageStatistics.date.asc()).all()
        
        # Toplam istatistikler
        total_results = db.query(Result).filter_by(tenant_id=tenant_id).count()
        total_views = db.query(Result).filter_by(tenant_id=tenant_id).with_entities(
            db.func.sum(Result.views_count)
        ).scalar() or 0
        total_forwards = db.query(Result).filter_by(tenant_id=tenant_id).with_entities(
            db.func.sum(Result.forwards_count)
        ).scalar() or 0
        
        # Kelime bazında istatistikler
        all_results = db.query(Result).filter_by(tenant_id=tenant_id).all()
        keyword_stats = {}
        for result in all_results:
            for keyword in (result.found_keywords or []):
                keyword_stats[keyword] = keyword_stats.get(keyword, 0) + 1
        
        return jsonify({
            'success': True,
            'daily_stats': [{
                'date': stat.date.isoformat(),
                'total_matches': stat.total_matches,
                'total_views': stat.total_views,
                'total_forwards': stat.total_forwards,
                'total_reactions': stat.total_reactions,
                'keyword_stats': stat.keyword_stats,
                'link_stats': stat.link_stats
            } for stat in daily_stats],
            'totals': {
                'total_results': total_results,
                'total_views': int(total_views),
                'total_forwards': int(total_forwards)
            },
            'keyword_stats': keyword_stats
        })
    finally:
        db.close()

@app.route('/api/admin/<int:tenant_id>/scan', methods=['POST'])
@login_required
@require_tenant_access('tenant_id')
def start_scan_api(tenant_id):
    """Tarama başlat"""
    try:
        config = get_tenant_config(tenant_id)
        if not config or not config.api_id or not config.get_api_hash():
            return jsonify({'success': False, 'message': 'API bilgileri eksik!'})
        
        if not config.group_ids:
            return jsonify({'success': False, 'message': 'Grup seçilmedi!'})
        
        # Tenant slug'ını al
        tenant = get_tenant(tenant_id)
        if not tenant:
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        tenant_slug = tenant.slug
        session_file = config.session_file_path or f'tenants/{tenant_slug}/session.session'
        if not os.path.exists(session_file):
            return jsonify({'success': False, 'message': 'Telegram girişi yapılmamış!'})
        
        # Bot zaten çalışıyor mu?
        if tenant_id in bot_statuses and bot_statuses[tenant_id].get('running'):
            return jsonify({'success': False, 'message': 'Bot zaten çalışıyor!'})
        
        # Botu başlat
        bot_process = subprocess.Popen(
            ['python', 'tg_monitor_tenant.py', str(tenant_id)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            universal_newlines=True
        )
        
        bot_processes[tenant_id] = bot_process
        bot_statuses[tenant_id] = {
            'running': True,
            'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        bot_logs[tenant_id] = []
        
        # Logları oku
        def read_logs():
            try:
                for line in iter(bot_process.stdout.readline, ''):
                    if line:
                        bot_logs[tenant_id].append(line.strip())
                        if len(bot_logs[tenant_id]) > 1000:
                            bot_logs[tenant_id] = bot_logs[tenant_id][-500:]
            except:
                pass
            finally:
                bot_statuses[tenant_id]['running'] = False
        
        threading.Thread(target=read_logs, daemon=True).start()
        
        return jsonify({'success': True, 'message': 'Tarama başlatıldı!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'})

@app.route('/api/admin/<int:tenant_id>/scan/status', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def get_scan_status_api(tenant_id):
    """Tarama durumunu al"""
    status = bot_statuses.get(tenant_id, {'running': False})
    logs = bot_logs.get(tenant_id, [])
    
    # Process kontrolü
    if tenant_id in bot_processes:
        process = bot_processes[tenant_id]
        try:
            poll_result = process.poll()
            if poll_result is not None:
                status['running'] = False
        except:
            status['running'] = False
    
    return jsonify({
        'success': True,
        'running': status.get('running', False),
        'start_time': status.get('start_time'),
        'logs': logs[-50:]  # Son 50 log
    })

# ==================== TELEGRAM ROUTES ====================

@app.route('/api/admin/<int:tenant_id>/telegram/groups', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def get_telegram_groups(tenant_id):
    """Telegram gruplarını listele"""
    try:
        client = get_telegram_client_for_tenant(tenant_id)
        if not client:
            return jsonify({'success': False, 'message': 'API bilgileri eksik!', 'groups': []})
        
        async def fetch_groups():
            try:
                await client.connect()
                if not await client.is_user_authorized():
                    raise Exception('Telegram girişi yapılmamış!')
                
                groups = []
                async for dialog in client.iter_dialogs():
                    if dialog.is_group or dialog.is_channel:
                        groups.append({
                            'id': dialog.id,
                            'name': dialog.name or 'İsimsiz Grup',
                            'unread': dialog.unread_count,
                            'is_channel': dialog.is_channel
                        })
                        if len(groups) >= 500:
                            break
                return groups
            finally:
                try:
                    await client.disconnect()
                except:
                    pass
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            groups = loop.run_until_complete(fetch_groups())
            return jsonify({'success': True, 'groups': groups})
        finally:
            loop.close()
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'groups': []})

@app.route('/api/admin/<int:tenant_id>/telegram/login', methods=['POST'])
@login_required
@require_tenant_access('tenant_id')
def telegram_login(tenant_id):
    """Telegram'a giriş yap"""
    try:
        data = request.json
        action = data.get('action')
        phone = data.get('phone', '').strip()
        
        config = get_tenant_config(tenant_id)
        if not config or not config.api_id or not config.get_api_hash():
            return jsonify({'success': False, 'message': 'API bilgileri eksik!'})
        
        # Tenant slug'ını al (session içinde)
        db = SessionLocal()
        try:
            tenant = db.query(Tenant).filter_by(id=tenant_id).first()
            if not tenant:
                return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
            tenant_slug = tenant.slug
        finally:
            db.close()
        
        session_path = config.session_file_path or f'tenants/{tenant_slug}/session.session'
        
        # Session dizinini oluştur
        session_dir = os.path.dirname(session_path)
        if session_dir and not os.path.exists(session_dir):
            os.makedirs(session_dir, exist_ok=True)
        
        # Session dosya adını düzelt (TelegramClient .session uzantısını ekler)
        # Session dosyası için mutlak yol kullan
        if not os.path.isabs(session_path):
            # Göreceli yol ise, çalışma dizinine göre mutlak yola çevir
            session_path = os.path.abspath(session_path)
        
        session_name = session_path.replace('.session', '')
        
        # Session dosyasının dizinini tekrar kontrol et ve oluştur
        session_dir = os.path.dirname(session_name)
        if session_dir and not os.path.exists(session_dir):
            os.makedirs(session_dir, exist_ok=True)
        
        # Session dosyası için izinleri kontrol et
        try:
            # Dizine yazma izni kontrolü
            if not os.access(session_dir, os.W_OK):
                logger.warning(f"   ⚠️  Session dizinine yazma izni yok: {session_dir}")
        except Exception as e:
            logger.warning(f"   ⚠️  Session dizin izni kontrolü hatası: {e}")
        
        client = TelegramClient(session_name, config.api_id, config.get_api_hash())
        
        async def handle_login():
            try:
                await client.connect()
                
                if await client.is_user_authorized():
                    await client.disconnect()
                    return {'success': True, 'message': 'Zaten giriş yapılmış!', 'requires_password': False}
                
                if action == 'send_code':
                    try:
                        sent_code = await client.send_code_request(phone)
                        client.session.save()
                        await client.disconnect()
                        return {'success': True, 'message': 'Kod gönderildi!'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        if 'PHONE_NUMBER_INVALID' in error_msg:
                            return {'success': False, 'message': 'Telefon numarası geçersiz!'}
                        elif 'FLOOD_WAIT' in error_msg:
                            return {'success': False, 'message': 'Çok fazla deneme! Lütfen bekleyin.'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                
                elif action == 'verify_code':
                    code = data.get('code', '').strip()
                    if not code:
                        return {'success': False, 'message': 'Kod gerekli!'}
                    
                    try:
                        result = await client.sign_in(phone, code)
                        client.session.save()
                        await client.disconnect()
                        
                        if os.path.exists(session_path):
                            return {'success': True, 'message': 'Giriş başarılı!', 'requires_password': False}
                        else:
                            return {'success': False, 'message': 'Session kaydedilemedi.'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        
                        if 'PASSWORD' in error_msg or 'SESSION_PASSWORD_NEEDED' in error_msg:
                            try:
                                client.session.save()
                            except:
                                pass
                            return {'success': True, 'message': 'İki faktörlü doğrulama gerekiyor', 'requires_password': True}
                        elif 'PHONE_CODE_INVALID' in error_msg:
                            return {'success': False, 'message': 'Kod geçersiz!'}
                        elif 'PHONE_CODE_EXPIRED' in error_msg:
                            return {'success': False, 'message': 'Kod süresi dolmuş!'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                
                elif action == 'verify_password':
                    password = data.get('password', '')
                    if not password:
                        return {'success': False, 'message': 'Şifre gerekli!'}
                    
                    try:
                        await client.sign_in(password=password)
                        client.session.save()
                        await client.disconnect()
                        
                        if os.path.exists(session_path):
                            return {'success': True, 'message': 'Giriş başarılı!'}
                        else:
                            return {'success': False, 'message': 'Session kaydedilemedi.'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        if 'PASSWORD' in error_msg:
                            return {'success': False, 'message': 'Şifre yanlış!'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                else:
                    return {'success': False, 'message': 'Geçersiz işlem!'}
            except Exception as e:
                error_msg = str(e)
                try:
                    await client.disconnect()
                except:
                    pass
                return {'success': False, 'message': f'Hata: {error_msg}'}
        
        # Event loop sorununu çöz - thread-safe event loop kullan
        try:
            loop = asyncio.get_event_loop()
            if loop.is_closed():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        
        try:
            result = loop.run_until_complete(handle_login())
        finally:
            # Loop'u kapatma, sadece temizle
            try:
                pending = asyncio.all_tasks(loop)
                for task in pending:
                    task.cancel()
            except:
                pass
        
        return jsonify(result)
    except Exception as e:
        logger.error(f"   ❌ Telegram login hatası: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Giriş hatası: {str(e)}'})

@app.route('/api/admin/<int:tenant_id>/telegram/groups/search', methods=['POST'])
@login_required
@require_tenant_access('tenant_id')
def search_telegram_groups(tenant_id):
    """Telegram'da grup ara"""
    try:
        data = request.json
        search_term = data.get('search_term', '').strip()
        
        if not search_term:
            return jsonify({'success': False, 'message': 'Arama terimi gerekli!', 'groups': []})
        
        client = get_telegram_client_for_tenant(tenant_id)
        if not client:
            return jsonify({'success': False, 'message': 'API bilgileri eksik!', 'groups': []})
        
        async def search_async():
            try:
                await client.connect()
                if not await client.is_user_authorized():
                    raise Exception('Telegram girişi yapılmamış!')
                
                groups = []
                search_lower = search_term.lower()
                
                async for dialog in client.iter_dialogs():
                    if dialog.is_group or dialog.is_channel:
                        dialog_name = (dialog.name or '').lower()
                        if search_lower in dialog_name:
                            groups.append({
                                'id': dialog.id,
                                'name': dialog.name or 'İsimsiz Grup',
                                'unread': dialog.unread_count,
                                'is_channel': dialog.is_channel
                            })
                            if len(groups) >= 50:
                                break
                return groups
            finally:
                try:
                    await client.disconnect()
                except:
                    pass
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            groups = loop.run_until_complete(search_async())
            return jsonify({'success': True, 'groups': groups})
        finally:
            loop.close()
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'groups': []})

@app.route('/api/admin/<int:tenant_id>/telegram/groups/add-by-username', methods=['POST'])
@login_required
@require_tenant_access('tenant_id')
def add_group_by_username(tenant_id):
    """Username'den grup ekle"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'success': False, 'message': 'Username gerekli!', 'group': None})
        
        if username.startswith('@'):
            username = username[1:]
        
        client = get_telegram_client_for_tenant(tenant_id)
        if not client:
            return jsonify({'success': False, 'message': 'API bilgileri eksik!', 'group': None})
        
        async def get_group_async():
            try:
                await client.connect()
                if not await client.is_user_authorized():
                    raise Exception('Telegram girişi yapılmamış!')
                
                entity = await client.get_entity(username)
                return {
                    'id': entity.id,
                    'name': getattr(entity, 'title', username) or username,
                    'is_channel': getattr(entity, 'broadcast', False),
                    'username': getattr(entity, 'username', username)
                }
            finally:
                try:
                    await client.disconnect()
                except:
                    pass
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            group = loop.run_until_complete(get_group_async())
            return jsonify({'success': True, 'group': group})
        finally:
            loop.close()
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'group': None})

@app.route('/api/admin/<int:tenant_id>/results/export', methods=['GET'])
@login_required
@require_tenant_access('tenant_id')
def export_results(tenant_id):
    """Sonuçları Excel formatında indir"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        db = SessionLocal()
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            
            query = db.query(Result).filter_by(tenant_id=tenant_id)
            
            if start_date:
                query = query.filter(Result.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                query = query.filter(Result.timestamp <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
            
            results = query.order_by(Result.timestamp.desc()).all()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Telegram Sonuçları"
            
            headers = ['Tarih', 'Grup', 'Grup ID', 'Bulunan Kelimeler', 'Bulunan Linkler', 
                      'Görüntülenme', 'Paylaşım', 'Reaksiyonlar', 'Yanıtlar', 'Mesaj İçeriği', 'Mesaj Linki']
            header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_num, result in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=result.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=2, value=result.group_name)
                ws.cell(row=row_num, column=3, value=result.group_id)
                ws.cell(row=row_num, column=4, value=', '.join(result.found_keywords or []))
                ws.cell(row=row_num, column=5, value=', '.join(result.found_links or []))
                ws.cell(row=row_num, column=6, value=result.views_count)
                ws.cell(row=row_num, column=7, value=result.forwards_count)
                ws.cell(row=row_num, column=8, value=str(result.reactions_detail or {}))
                ws.cell(row=row_num, column=9, value=result.replies_count)
                ws.cell(row=row_num, column=10, value=result.message_text)
                ws.cell(row=row_num, column=11, value=result.message_link)
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 50
            ws.column_dimensions['K'].width = 40
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            from flask import Response
            filename = f"telegram_sonuclari_{start_date or 'tum'}_{end_date or 'tum'}.xlsx"
            return Response(
                output.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        finally:
            db.close()
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl gerekli!'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# ==================== LEGACY API ROUTES (index.html için) ====================
# Bu route'lar eski index.html ile uyumluluk için
# Tenant ID otomatik olarak belirlenir

@app.route('/api/config', methods=['GET'])
@login_required
def get_config_api_legacy():
    """Config'i getir (eski format)"""
    try:
        logger.info("📥 GET /api/config çağrıldı")
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı! Lütfen önce bir grup oluşturun.'})
        
        config = get_tenant_config(tenant_id)
        if not config:
            logger.warning(f"   ⚠️  Config bulunamadı (tenant_id: {tenant_id})")
            return jsonify({'success': False, 'message': 'Config bulunamadı!'})
        
        logger.info("   ✅ Config başarıyla alındı")
        # Eski format
        return jsonify({
            'API_ID': config.api_id or '',
            'API_HASH': '***' if config.api_hash_encrypted else '',
            'PHONE_NUMBER': config.phone_number or '',
            'GROUP_IDS': config.group_ids or [],
            'SEARCH_KEYWORDS': config.search_keywords or [],
            'SEARCH_LINKS': config.search_links or [],
            'SCAN_TIME_RANGE': config.scan_time_range or '7days'
        })
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/config', methods=['POST'])
@login_required
def save_config_api_legacy():
    """Config'i kaydet (eski format)"""
    try:
        logger.info("📥 POST /api/config çağrıldı")
        
        if not request.is_json:
            logger.error("   ❌ Request JSON değil!")
            return jsonify({'success': False, 'message': 'Request JSON formatında olmalı!'}), 400
        
        data = request.json
        logger.info(f"   Request Data: {json.dumps(data, indent=2, ensure_ascii=False)}")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı! Lütfen önce bir grup oluşturun.'})
        
        update_data = {}
        
        if 'API_ID' in data:
            update_data['api_id'] = data['API_ID']
        if 'API_HASH' in data and data['API_HASH'] != '***':
            update_data['api_hash'] = data['API_HASH']
        if 'PHONE_NUMBER' in data:
            update_data['phone_number'] = data['PHONE_NUMBER']
        if 'GROUP_IDS' in data:
            update_data['group_ids'] = data['GROUP_IDS']
        if 'SEARCH_KEYWORDS' in data:
            update_data['search_keywords'] = [kw.strip() for kw in data['SEARCH_KEYWORDS'] if kw.strip()]
        if 'SEARCH_LINKS' in data:
            update_data['search_links'] = [link.strip() for link in data['SEARCH_LINKS'] if link.strip()]
        if 'SCAN_TIME_RANGE' in data:
            update_data['scan_time_range'] = data['SCAN_TIME_RANGE']
        
        logger.info(f"   Update Data: {update_data}")
        
        config = update_tenant_config(tenant_id, **update_data)
        if config:
            logger.info("   ✅ Config başarıyla güncellendi")
            return jsonify({'success': True, 'message': 'Ayarlar kaydedildi!'})
        else:
            logger.warning("   ⚠️  Config bulunamadı veya güncellenemedi")
            return jsonify({'success': False, 'message': 'Config bulunamadı!'})
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/groups', methods=['GET'])
@login_required
def get_groups_legacy():
    """Telegram gruplarını listele (eski format)"""
    try:
        logger.info("📥 GET /api/groups çağrıldı")
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!', 'groups': []})
        
        return get_telegram_groups(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'groups': []}), 500

@app.route('/api/groups/search', methods=['POST'])
@login_required
def search_groups_legacy():
    """Telegram'da grup ara (eski format)"""
    try:
        logger.info("📥 POST /api/groups/search çağrıldı")
        
        if not request.is_json:
            logger.error("   ❌ Request JSON değil!")
            return jsonify({'success': False, 'message': 'Request JSON formatında olmalı!', 'groups': []}), 400
        
        data = request.json
        logger.info(f"   Request Data: {json.dumps(data, indent=2, ensure_ascii=False)}")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!', 'groups': []})
        
        return search_telegram_groups(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'groups': []}), 500

@app.route('/api/groups/add-by-username', methods=['POST'])
@login_required
def add_group_by_username_legacy():
    """Username'den grup ekle (eski format)"""
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        return jsonify({'success': False, 'message': 'Tenant bulunamadı!', 'group': None})
    
    return add_group_by_username(tenant_id)

@app.route('/api/results', methods=['GET'])
@login_required
def get_results_legacy():
    """Sonuçları al (eski format)"""
    try:
        logger.info("📥 GET /api/results çağrıldı")
        logger.info(f"   Query Params: {dict(request.args)}")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            # Eğer süper admin ise, ilk tenant'ı kullan
            if current_user.is_super_admin:
                db = SessionLocal()
                try:
                    first_tenant = db.query(Tenant).filter_by(is_active=True).first()
                    if first_tenant:
                        tenant_id = first_tenant.id
                        logger.info(f"   Süper admin için ilk tenant kullanılıyor: {tenant_id}")
                    else:
                        return jsonify({'success': False, 'message': 'Hiç aktif grup yok!', 'results': []})
                finally:
                    db.close()
            else:
                return jsonify({'success': False, 'message': 'Tenant bulunamadı! Lütfen giriş yaparken bir grup seçin.', 'results': []})
        
        return get_results_api(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'results': []}), 500

@app.route('/api/results/clear', methods=['POST'])
@login_required
def clear_results_legacy():
    """Sonuçları temizle (eski format)"""
    try:
        logger.info("📥 POST /api/results/clear çağrıldı")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        
        db = SessionLocal()
        try:
            deleted_count = db.query(Result).filter_by(tenant_id=tenant_id).delete()
            db.commit()
            logger.info(f"   ✅ {deleted_count} sonuç silindi")
            return jsonify({'success': True, 'message': 'Sonuçlar temizlendi!'})
        except Exception as e:
            db.rollback()
            logger.error(f"   ❌ Database hatası: {str(e)}")
            logger.error(f"   Traceback: {traceback.format_exc()}")
            return jsonify({'success': False, 'message': f'Hata: {str(e)}'})
        finally:
            db.close()
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/telegram-login', methods=['POST'])
@login_required
def telegram_login_legacy():
    """Telegram'a giriş yap (eski format)"""
    try:
        logger.info("📥 POST /api/telegram-login çağrıldı")
        
        if not request.is_json:
            logger.error("   ❌ Request JSON değil!")
            return jsonify({'success': False, 'message': 'Request JSON formatında olmalı!'}), 400
        
        data = request.json
        logger.info(f"   Request Data: {json.dumps(data, indent=2, ensure_ascii=False)}")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        
        return telegram_login(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/scan', methods=['POST'])
@login_required
def start_scan_legacy():
    """Tarama başlat (eski format)"""
    try:
        logger.info("📥 POST /api/scan çağrıldı")
        
        if request.is_json:
            logger.info(f"   Request Data: {json.dumps(request.json, indent=2, ensure_ascii=False)}")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        
        return start_scan_api(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/scan-status', methods=['GET'])
@login_required
def get_scan_status_legacy():
    """Tarama durumunu al (eski format)"""
    try:
        logger.info("📥 GET /api/scan-status çağrıldı")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        
        return get_scan_status_api(tenant_id)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/test-telegram', methods=['POST'])
@login_required
def test_telegram_legacy():
    """Telegram API testi (eski format)"""
    try:
        logger.info("📥 POST /api/test-telegram çağrıldı")
        
        tenant_id = get_current_tenant_id()
        logger.info(f"   Tenant ID: {tenant_id}")
        
        if not tenant_id:
            logger.warning("   ⚠️  Tenant bulunamadı!")
            return jsonify({'success': False, 'message': 'Tenant bulunamadı!'})
        
        client = get_telegram_client_for_tenant(tenant_id)
        if not client:
            logger.warning("   ⚠️  Telegram client oluşturulamadı (API bilgileri eksik)")
            return jsonify({'success': False, 'message': 'API bilgileri eksik! Lütfen Ayarlar sekmesinden API ID ve API Hash bilgilerinizi girin.'})
        
        async def test():
            try:
                await client.connect()
                if await client.is_user_authorized():
                    await client.disconnect()
                    logger.info("   ✅ Telegram bağlantısı başarılı")
                    return {'success': True, 'message': 'Telegram bağlantısı başarılı!'}
                else:
                    await client.disconnect()
                    logger.warning("   ⚠️  Telegram girişi yapılmamış")
                    return {'success': False, 'message': 'Telegram girişi yapılmamış!'}
            except Exception as e:
                try:
                    await client.disconnect()
                except:
                    pass
                logger.error(f"   ❌ Telegram test hatası: {str(e)}")
                return {'success': False, 'message': f'Hata: {str(e)}'}
        
        # Event loop sorununu çöz - thread-safe event loop kullan
        try:
            loop = asyncio.get_event_loop()
            if loop.is_closed():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        
        try:
            result = loop.run_until_complete(test())
        finally:
            # Loop'u kapatma, sadece temizle
            try:
                pending = asyncio.all_tasks(loop)
                for task in pending:
                    task.cancel()
            except:
                pass
        
        return jsonify(result)
    except Exception as e:
        logger.error(f"   ❌ Hata: {str(e)}")
        logger.error(f"   Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

if __name__ == '__main__':
    # Database'i başlat
    print("🔧 Database başlatılıyor...")
    try:
        init_db()
        create_super_admin()
        
        # Hazır grupları oluştur
        print("🔧 Hazır gruplar oluşturuluyor...")
        db = SessionLocal()
        try:
            # Süper admin kullanıcısını bul
            super_admin = db.query(User).filter_by(role='super_admin').first()
            if super_admin:
                groups = ['Gala', 'Hit', 'Pipo', 'Office', 'Padisah']
                created_count = 0
                for group_name in groups:
                    # Grup zaten var mı kontrol et
                    existing = db.query(Tenant).filter_by(name=group_name).first()
                    if not existing:
                        try:
                            tenant = create_tenant(group_name, super_admin.id)
                            print(f"✅ '{group_name}' grubu oluşturuldu (ID: {tenant.id})")
                            created_count += 1
                        except Exception as e:
                            print(f"⚠️  '{group_name}' grubu oluşturulamadı: {e}")
                    else:
                        print(f"ℹ️  '{group_name}' grubu zaten mevcut (ID: {existing.id})")
                print(f"🎉 {created_count} yeni grup oluşturuldu!")
            else:
                print("⚠️  Süper admin bulunamadı, gruplar oluşturulamadı!")
        finally:
            db.close()
        
        print("✅ Database hazır!")
        print("🔐 Süper Admin: padisah_admin / P@d1$@h2024!Secure#Admin")
    except Exception as e:
        print(f"⚠️  Database hatası (devam ediliyor): {e}")
        import traceback
        traceback.print_exc()
    
    # Port'u environment variable'dan al (Coolify PORT kullanır)
    port = int(os.environ.get('PORT', 5000))
    print("🌐 Web paneli başlatılıyor...")
    print(f"📱 Port: {port}")
    print(f"🌍 Host: 0.0.0.0")
    app.run(debug=False, host='0.0.0.0', port=port)

