"""
Telegram Monitoring Bot - Web Panel
Flask tabanlı web arayüzü
"""

from flask import Flask, render_template, request, jsonify, send_from_directory, session, redirect, url_for
from flask_cors import CORS
import asyncio
import json
from config_manager import load_config, save_config, get_config
from telethon import TelegramClient
import os
from functools import wraps
from hashlib import sha256

app = Flask(__name__, static_folder='.')
app.secret_key = os.environ.get('SECRET_KEY', 'padisah-telegram-monitoring-secret-key-change-in-production')
CORS(app)

# Admin bilgileri
ADMIN_USERNAME = 'oyku_admin'
ADMIN_PASSWORD_HASH = sha256('Havuclukek'.encode()).hexdigest()

# Global Telegram client (lazy loading)
telegram_client = None

def get_telegram_client():
    """Telegram client'ı al veya oluştur"""
    global telegram_client
    if telegram_client is None:
        config = load_config()
        if config['API_ID'] and config['API_HASH']:
            telegram_client = TelegramClient('session', config['API_ID'], config['API_HASH'])
    return telegram_client

def login_required(f):
    """Login kontrolü decorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login sayfası"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        password_hash = sha256(password.encode()).hexdigest()
        
        if username == ADMIN_USERNAME and password_hash == ADMIN_PASSWORD_HASH:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Kullanıcı adı veya şifre hatalı!')
    
    # GET isteği - login sayfasını göster
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Çıkış yap"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """Ana sayfa"""
    return render_template('index.html')

@app.route('/s-l400.jpg')
def serve_logo():
    """Logo dosyasını serve et"""
    return send_from_directory('.', 's-l400.jpg')

@app.route('/logoSeffaf.png')
def serve_logo_seffaf():
    """Şeffaf logo dosyasını serve et"""
    return send_from_directory('.', 'logoSeffaf.png')

@app.route('/api/config', methods=['GET'])
@login_required
def get_config_api():
    """Config'i getir"""
    config = load_config()
    # API_HASH'i güvenlik için gizle
    safe_config = config.copy()
    if safe_config.get('API_HASH'):
        safe_config['API_HASH'] = '***' if safe_config['API_HASH'] else ''
    return jsonify(safe_config)

@app.route('/api/config', methods=['POST'])
@login_required
def save_config_api():
    """Config'i kaydet"""
    try:
        data = request.json
        
        # Mevcut config'i yükle
        current_config = load_config()
        
        # Yeni değerleri güncelle
        if 'API_ID' in data:
            current_config['API_ID'] = data['API_ID']
        if 'API_HASH' in data:
            # Eğer *** ise değiştirme (güvenlik)
            if data['API_HASH'] != '***':
                current_config['API_HASH'] = data['API_HASH']
        if 'PHONE_NUMBER' in data:
            current_config['PHONE_NUMBER'] = data['PHONE_NUMBER']
        if 'GROUP_IDS' in data:
            # GROUP_IDS artık obje listesi olabilir: [{id, name, startDate, endDate}, ...]
            group_ids = []
            for item in data['GROUP_IDS']:
                if isinstance(item, dict):
                    # Obje formatında: {id, name, startDate, endDate}
                    group_obj = {
                        'id': item.get('id'),
                        'name': item.get('name'),  # İsmi de kaydet
                        'startDate': item.get('startDate') or None,
                        'endDate': item.get('endDate') or None
                    }
                    # Sadece id varsa ekle
                    if group_obj['id']:
                        group_ids.append(group_obj)
                elif isinstance(item, (int, str)):
                    # Eski format: sadece ID
                    try:
                        group_ids.append(int(item))
                    except:
                        pass
            current_config['GROUP_IDS'] = group_ids
        if 'SEARCH_KEYWORDS' in data:
            current_config['SEARCH_KEYWORDS'] = [kw.strip() for kw in data['SEARCH_KEYWORDS'] if kw.strip()]
        if 'SEARCH_LINKS' in data:
            current_config['SEARCH_LINKS'] = [link.strip() for link in data['SEARCH_LINKS'] if link.strip()]
        if 'RESULTS_FILE' in data:
            current_config['RESULTS_FILE'] = data['RESULTS_FILE']
        if 'SCAN_TIME_RANGE' in data:
            current_config['SCAN_TIME_RANGE'] = data['SCAN_TIME_RANGE']
        
        # Kaydet
        if save_config(current_config):
            return jsonify({'success': True, 'message': 'Ayarlar başarıyla kaydedildi!'})
        else:
            return jsonify({'success': False, 'message': 'Ayarlar kaydedilirken hata oluştu!'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/groups', methods=['GET'])
@login_required
def get_groups():
    """Telegram gruplarını listele"""
    try:
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            return jsonify({'success': False, 'message': 'API bilgileri eksik! Lütfen Ayarlar sekmesinden API ID ve API Hash girin.', 'groups': []})
        
        # Async fonksiyonu çalıştır
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            groups = loop.run_until_complete(fetch_groups())
        finally:
            loop.close()
        
        if not groups:
            return jsonify({'success': False, 'message': 'Grup bulunamadı veya Telegram\'a bağlanılamadı. Lütfen önce tg_monitor.py dosyasını çalıştırıp giriş yapın.', 'groups': []})
        
        return jsonify({'success': True, 'groups': groups})
    except Exception as e:
        error_msg = str(e)
        return jsonify({'success': False, 'message': error_msg, 'groups': []})

@app.route('/api/groups/search', methods=['POST'])
@login_required
def search_groups():
    """Telegram'da grup ara (kullanıcının üye olduğu gruplar içinde)"""
    try:
        data = request.json
        search_term = data.get('search_term', '').strip()
        
        if not search_term:
            return jsonify({'success': False, 'message': 'Arama terimi gerekli!', 'groups': []})
        
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            return jsonify({'success': False, 'message': 'API bilgileri eksik!', 'groups': []})
        
        # Async fonksiyonu çalıştır
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            groups = loop.run_until_complete(search_groups_async(search_term))
        finally:
            loop.close()
        
        return jsonify({'success': True, 'groups': groups})
    except Exception as e:
        error_msg = str(e)
        print(f"search_groups endpoint hatası: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Arama hatası: {error_msg}', 'groups': []})

@app.route('/api/groups/add-by-username', methods=['POST'])
@login_required
def add_group_by_username():
    """Username'den grup ekle"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'success': False, 'message': 'Username gerekli!', 'group': None})
        
        # @ işaretini kaldır
        if username.startswith('@'):
            username = username[1:]
        
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            return jsonify({'success': False, 'message': 'API bilgileri eksik!', 'group': None})
        
        # Async fonksiyonu çalıştır
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            group = loop.run_until_complete(get_group_by_username_async(username))
        finally:
            loop.close()
        
        return jsonify({'success': True, 'group': group})
    except Exception as e:
        error_msg = str(e)
        print(f"add_group_by_username hatası: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': error_msg, 'group': None})

async def search_groups_async(search_term):
    """Telegram'da grup ara (async)"""
    try:
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            raise Exception('API bilgileri eksik!')
        
        # Session dosyası kontrolü
        session_file = 'session.session'
        if not os.path.exists(session_file):
            raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.')
        
        # Yeni client oluştur
        client = TelegramClient('session', config['API_ID'], config['API_HASH'])
        
        try:
            await client.connect()
            
            if not await client.is_user_authorized():
                raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.')
            
            # Kullanıcının üye olduğu tüm grupları çek ve arama terimine göre filtrele
            groups = []
            search_lower = search_term.lower()
            
            async for dialog in client.iter_dialogs():
                if dialog.is_group or dialog.is_channel:
                    # Grup/kanal adında arama terimi var mı kontrol et
                    dialog_name = (dialog.name or '').lower()
                    if search_lower in dialog_name:
                        groups.append({
                            'id': dialog.id,
                            'name': dialog.name or 'İsimsiz Grup',
                            'unread': dialog.unread_count,
                            'is_channel': dialog.is_channel
                        })
                        if len(groups) >= 50:  # Maksimum 50 sonuç
                            break
            
            return groups
            
        finally:
            try:
                await client.disconnect()
            except:
                pass
                
    except Exception as e:
        error_msg = str(e)
        print(f"search_groups_async hatası: {error_msg}")
        raise Exception(f'Grup arama hatası: {error_msg}')

async def get_group_by_username_async(username):
    """Username'den grup bilgilerini çek"""
    try:
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            raise Exception('API bilgileri eksik!')
        
        # Session dosyası kontrolü
        session_file = 'session.session'
        if not os.path.exists(session_file):
            raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.')
        
        client = TelegramClient('session', config['API_ID'], config['API_HASH'])
        
        try:
            await client.connect()
            
            if not await client.is_user_authorized():
                raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.')
            
            # Username'den entity'yi al
            try:
                entity = await client.get_entity(username)
                return {
                    'id': entity.id,
                    'name': getattr(entity, 'title', username) or username,
                    'is_channel': getattr(entity, 'broadcast', False),
                    'username': getattr(entity, 'username', username)
                }
            except Exception as e:
                error_msg = str(e)
                if 'not found' in error_msg.lower() or 'could not find' in error_msg.lower():
                    raise Exception(f'Grup/kanal bulunamadı: @{username} - Bu grup/kanal mevcut değil veya erişim izniniz yok.')
                else:
                    raise Exception(f'Grup bulunamadı: {error_msg}')
            
        finally:
            try:
                await client.disconnect()
            except:
                pass
                
    except Exception as e:
        error_msg = str(e)
        print(f"get_group_by_username_async hatası: {error_msg}")
        raise Exception(f'Grup çekme hatası: {error_msg}')

async def fetch_groups():
    """Telegram gruplarını çek"""
    try:
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            raise Exception('API bilgileri eksik! Lütfen API ID ve API Hash girin.')
        
        # Yeni client oluştur (her seferinde fresh connection)
        client = TelegramClient('session', config['API_ID'], config['API_HASH'])
        
        try:
            # Bağlan
            await client.connect()
            
            # Kullanıcı yetkilendirilmiş mi kontrol et
            if not await client.is_user_authorized():
                raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce botu çalıştırıp giriş yapın (tg_monitor.py).')
            
            # Grupları çek
            groups = []
            count = 0
            async for dialog in client.iter_dialogs():
                if dialog.is_group:
                    groups.append({
                        'id': dialog.id,
                        'name': dialog.name or 'İsimsiz Grup',
                        'unread': dialog.unread_count
                    })
                    count += 1
                    if count > 500:  # Limit (çok fazla grup varsa)
                        break
            
            return groups
            
        finally:
            # Bağlantıyı kapat
            try:
                await client.disconnect()
            except:
                pass
                
    except Exception as e:
        error_msg = str(e)
        print(f"Grup çekme hatası: {error_msg}")
        # Daha açıklayıcı hata mesajları
        if 'API_ID_INVALID' in error_msg or 'api_id' in error_msg.lower():
            raise Exception('API ID geçersiz! Lütfen doğru API ID girin.')
        elif 'API_HASH_INVALID' in error_msg or 'api_hash' in error_msg.lower():
            raise Exception('API Hash geçersiz! Lütfen doğru API Hash girin.')
        elif 'not authorized' in error_msg.lower() or 'unauthorized' in error_msg.lower():
            raise Exception('Telegram hesabınıza giriş yapılmamış! Lütfen önce tg_monitor.py dosyasını çalıştırıp giriş yapın.')
        elif 'connection' in error_msg.lower() or 'timeout' in error_msg.lower():
            raise Exception('Telegram\'a bağlanılamadı! İnternet bağlantınızı kontrol edin.')
        else:
            raise Exception(f'Grup yükleme hatası: {error_msg}')

@app.route('/api/results', methods=['GET'])
@login_required
def get_results():
    """Sonuçları getir (tarih filtresi ile)"""
    try:
        config = load_config()
        results_file = config.get('RESULTS_FILE', 'results.txt')
        
        # Tarih filtresi parametrelerini al
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        
        if not os.path.exists(results_file):
            return jsonify({'success': True, 'results': []})
        
        # Sonuçları oku (basit format)
        with open(results_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Sonuçları parse et (basit format için)
        results = []
        sections = content.split('=' * 60)
        for section in sections:
            if section.strip():
                result = {}
                lines = section.strip().split('\n')
                current_key = None
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    if ':' in line and not line.startswith(' '):
                        # Yeni bir key-value çifti
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            current_key = parts[0].strip()
                            value = parts[1].strip()
                            # Eğer key zaten varsa, yeni değeri ekle
                            if current_key in result:
                                if current_key == 'Mesaj':
                                    result[current_key] = result[current_key] + '\n' + value
                                else:
                                    result[current_key] = result[current_key] + ' ' + value
                            else:
                                result[current_key] = value
                    elif current_key and line:
                        # Önceki key'in devamı (çok satırlı mesaj için)
                        if current_key == 'Mesaj':
                            result[current_key] = (result.get(current_key, '') + '\n' + line).strip()
                        else:
                            result[current_key] = (result.get(current_key, '') + ' ' + line).strip()
                
                # Mesaj içeriğini düzelt (eğer boşsa veya sadece boşluk varsa)
                if 'Mesaj' in result:
                    result['Mesaj'] = result['Mesaj'].strip()
                    if not result['Mesaj'] or result['Mesaj'] == 'N/A':
                        result['Mesaj'] = 'Mesaj içeriği bulunamadı.'
                
                if result:
                    results.append(result)
        
        # Tarih filtresi uygula
        if start_date or end_date:
            from datetime import datetime
            filtered_results = []
            for result in results:
                timestamp = result.get('Tarih', result.get('timestamp', ''))
                if not timestamp:
                    continue
                
                try:
                    # Tarihi parse et
                    date_str = timestamp.split(' ')[0]  # Sadece tarih kısmı
                    result_date = datetime.strptime(date_str, '%Y-%m-%d')
                    
                    # Başlangıç tarihi kontrolü
                    if start_date:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        if result_date < start_dt:
                            continue
                    
                    # Bitiş tarihi kontrolü
                    if end_date:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        if result_date > end_dt:
                            continue
                    
                    filtered_results.append(result)
                except:
                    # Tarih parse edilemezse atla
                    continue
            
            results = filtered_results
        
        # Tüm sonuçları döndür (limit yok)
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'results': []})

@app.route('/api/results/clear', methods=['POST'])
@login_required
def clear_results():
    """Eski sonuçları temizle"""
    try:
        config = load_config()
        results_file = config.get('RESULTS_FILE', 'results.txt')
        
        if os.path.exists(results_file):
            # Dosyayı tamamen sil veya boş yap
            try:
                # Önce dosyayı boş yap
                with open(results_file, 'w', encoding='utf-8') as f:
                    f.write('')
                # Dosyayı kapat ve kontrol et
                import time
                time.sleep(0.1)  # Kısa bir bekleme
                
                # Dosyanın gerçekten boş olduğunu kontrol et
                with open(results_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    if content.strip():
                        # Hala içerik varsa tekrar dene
                        with open(results_file, 'w', encoding='utf-8') as f2:
                            f2.write('')
                
                return jsonify({'success': True, 'message': 'Tüm sonuçlar başarıyla temizlendi!'})
            except Exception as e:
                return jsonify({'success': False, 'message': f'Dosya temizleme hatası: {str(e)}'})
        else:
            return jsonify({'success': True, 'message': 'Temizlenecek sonuç dosyası bulunamadı.'})
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'details': error_details})

@app.route('/api/results/export', methods=['GET'])
@login_required
def export_results():
    """Sonuçları Excel formatında indir"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        config = load_config()
        results_file = config.get('RESULTS_FILE', 'results.txt')
        
        # Tarih filtresi parametrelerini al
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        
        if not os.path.exists(results_file):
            return jsonify({'success': False, 'message': 'Sonuç dosyası bulunamadı!'}), 404
        
        # Sonuçları oku ve parse et
        with open(results_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        results = []
        sections = content.split('=' * 60)
        for section in sections:
            if section.strip():
                result = {}
                lines = section.strip().split('\n')
                current_key = None
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    if ':' in line and not line.startswith(' '):
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            current_key = parts[0].strip()
                            value = parts[1].strip()
                            if current_key in result:
                                if current_key == 'Mesaj':
                                    result[current_key] = result[current_key] + '\n' + value
                                else:
                                    result[current_key] = result[current_key] + ' ' + value
                            else:
                                result[current_key] = value
                    elif current_key and line:
                        if current_key == 'Mesaj':
                            result[current_key] = (result.get(current_key, '') + '\n' + line).strip()
                        else:
                            result[current_key] = (result.get(current_key, '') + ' ' + line).strip()
                
                if 'Mesaj' in result:
                    result['Mesaj'] = result['Mesaj'].strip()
                
                if result:
                    results.append(result)
        
        # Tarih filtresi uygula
        if start_date or end_date:
            from datetime import datetime
            filtered_results = []
            for result in results:
                timestamp = result.get('Tarih', result.get('timestamp', ''))
                if not timestamp:
                    continue
                
                try:
                    date_str = timestamp.split(' ')[0]
                    result_date = datetime.strptime(date_str, '%Y-%m-%d')
                    
                    if start_date:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        if result_date < start_dt:
                            continue
                    
                    if end_date:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        if result_date > end_dt:
                            continue
                    
                    filtered_results.append(result)
                except:
                    continue
            
            results = filtered_results
        
        # Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Telegram Sonuçları"
        
        # Başlık satırı
        headers = ['Tarih', 'Grup', 'Grup ID', 'Bulunan Kelimeler', 'Bulunan Linkler', 'Mesaj İçeriği', 'Mesaj Linki']
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Veri satırları
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=result.get('Tarih', result.get('timestamp', '')))
            ws.cell(row=row_num, column=2, value=result.get('Grup', result.get('group_name', '')))
            ws.cell(row=row_num, column=3, value=result.get('group_id', ''))
            
            # Bulunan kelimeler
            found_keywords = result.get('Bulunan Kelimeler', '')
            if not found_keywords and result.get('found_keywords'):
                found_keywords = ', '.join(result['found_keywords']) if isinstance(result['found_keywords'], list) else str(result['found_keywords'])
            ws.cell(row=row_num, column=4, value=found_keywords)
            
            # Bulunan linkler
            found_links = result.get('Bulunan Linkler', '')
            if not found_links and result.get('found_links'):
                found_links = ', '.join(result['found_links']) if isinstance(result['found_links'], list) else str(result['found_links'])
            ws.cell(row=row_num, column=5, value=found_links)
            
            # Mesaj içeriği
            message_text = result.get('Mesaj', result.get('message_text', ''))
            ws.cell(row=row_num, column=6, value=message_text)
            
            # Mesaj linki
            message_link = result.get('Link', result.get('message_link', ''))
            ws.cell(row=row_num, column=7, value=message_link)
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 40
        
        # Excel dosyasını memory'de oluştur
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Response oluştur
        from flask import Response
        filename = f"telegram_sonuclari_{start_date or 'tum'}_{end_date or 'tum'}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        return jsonify({'success': False, 'message': 'Excel export için openpyxl kütüphanesi gerekli! Lütfen: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

@app.route('/api/archive', methods=['GET'])
@login_required
def get_archive():
    """Arşiv sonuçlarını getir (tüm sonuçlar)"""
    try:
        config = load_config()
        results_file = config.get('RESULTS_FILE', 'results.txt')
        
        # Tarih filtresi parametrelerini al
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        
        if not os.path.exists(results_file):
            return jsonify({'success': True, 'results': []})
        
        # Sonuçları oku (basit format)
        with open(results_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Sonuçları parse et (basit format için)
        results = []
        sections = content.split('=' * 60)
        for section in sections:
            if section.strip():
                result = {}
                lines = section.strip().split('\n')
                current_key = None
                for line in lines:
                    if ':' in line and not line.startswith(' '):
                        # Yeni bir key-value çifti
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            current_key = parts[0].strip()
                            result[current_key] = parts[1].strip()
                    elif current_key and line.strip():
                        # Önceki key'in devamı (çok satırlı mesaj için)
                        if current_key == 'Mesaj':
                            result[current_key] = (result.get(current_key, '') + '\n' + line.strip()).strip()
                        else:
                            result[current_key] = (result.get(current_key, '') + ' ' + line.strip()).strip()
                if result:
                    results.append(result)
        
        # Tarih filtresi uygula
        if start_date or end_date:
            from datetime import datetime
            filtered_results = []
            for result in results:
                timestamp = result.get('Tarih', result.get('timestamp', ''))
                if not timestamp:
                    continue
                
                try:
                    # Tarihi parse et
                    date_str = timestamp.split(' ')[0]  # Sadece tarih kısmı
                    result_date = datetime.strptime(date_str, '%Y-%m-%d')
                    
                    # Başlangıç tarihi kontrolü
                    if start_date:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        if result_date < start_dt:
                            continue
                    
                    # Bitiş tarihi kontrolü
                    if end_date:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        if result_date > end_dt:
                            continue
                    
                    filtered_results.append(result)
                except:
                    # Tarih parse edilemezse atla
                    continue
            
            results = filtered_results
        
        # Tüm sonuçlar (ters sırada - en yeniden eskiye)
        return jsonify({'success': True, 'results': list(reversed(results))})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}', 'results': []})

@app.route('/api/test-telegram', methods=['POST'])
@login_required
def test_telegram_api():
    """Telegram API bilgilerini test et"""
    import time
    import uuid
    
    try:
        data = request.json
        api_id = data.get('API_ID', '').strip()
        api_hash = data.get('API_HASH', '').strip()
        phone_number = data.get('PHONE_NUMBER', '').strip()
        
        if not api_id or not api_hash:
            return jsonify({'success': False, 'message': 'API ID ve API Hash gereklidir!'})
        
        # Her test için benzersiz session dosyası kullan (database lock önlemek için)
        test_session_name = f'test_session_{uuid.uuid4().hex[:8]}_{int(time.time())}'
        test_client = None
        
        # Async test fonksiyonu
        async def test_connection():
            nonlocal test_client
            try:
                # Test için benzersiz session dosyası oluştur
                test_client = TelegramClient(test_session_name, api_id, api_hash)
                
                # Kısa timeout ile bağlan (5 saniye)
                await asyncio.wait_for(test_client.connect(), timeout=5.0)
                
                # Bağlantı başarılı mı kontrol et
                if not await test_client.is_user_authorized():
                    # Kullanıcı yetkilendirilmemiş, bu normal (ilk bağlantı)
                    # Sadece API bilgilerinin doğru olup olmadığını kontrol ediyoruz
                    return {'success': True, 'message': 'API bilgileri doğru! İlk bağlantı için telefon numaranıza kod gönderilecek.'}
                else:
                    # Zaten yetkilendirilmiş
                    user = await test_client.get_me()
                    return {'success': True, 'message': f'API bağlantısı başarılı! Hesap: {user.first_name or "Bilinmeyen"}'}
            except asyncio.TimeoutError:
                return {'success': False, 'message': 'Bağlantı zaman aşımı! Lütfen tekrar deneyin.'}
            except Exception as e:
                error_msg = str(e)
                if 'database is locked' in error_msg.lower():
                    return {'success': False, 'message': 'Session dosyası kullanımda! Lütfen birkaç saniye bekleyip tekrar deneyin.'}
                elif 'PHONE_NUMBER_INVALID' in error_msg or 'invalid phone' in error_msg.lower():
                    return {'success': False, 'message': 'Telefon numarası geçersiz! Format: +90XXXXXXXXXX'}
                elif 'API_ID_INVALID' in error_msg or 'api_id' in error_msg.lower():
                    return {'success': False, 'message': 'API ID geçersiz!'}
                elif 'API_HASH_INVALID' in error_msg or 'api_hash' in error_msg.lower():
                    return {'success': False, 'message': 'API Hash geçersiz!'}
                elif 'FLOOD_WAIT' in error_msg:
                    return {'success': False, 'message': 'Çok fazla istek! Lütfen birkaç dakika bekleyin.'}
                else:
                    return {'success': False, 'message': f'Bağlantı hatası: {error_msg}'}
            finally:
                # Session dosyasını temizle
                try:
                    if test_client:
                        await test_client.disconnect()
                        await test_client.log_out()
                except:
                    pass
                finally:
                    # Session dosyalarını sil
                    try:
                        import os
                        session_files = [
                            f'{test_session_name}.session',
                            f'{test_session_name}.session-journal'
                        ]
                        for session_file in session_files:
                            if os.path.exists(session_file):
                                try:
                                    os.remove(session_file)
                                except:
                                    pass
                    except:
                        pass
        
        # Async fonksiyonu çalıştır
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(test_connection())
        finally:
            loop.close()
        
        return jsonify(result)
        
    except Exception as e:
        error_msg = str(e)
        if 'database is locked' in error_msg.lower():
            return jsonify({'success': False, 'message': 'Session dosyası kullanımda! Lütfen birkaç saniye bekleyip tekrar deneyin.'})
        return jsonify({'success': False, 'message': f'Test hatası: {error_msg}'})

# Basitleştirilmiş giriş - phone_code_hash session'da otomatik saklanır
@app.route('/api/telegram-login', methods=['POST'])
@login_required
def telegram_login():
    """Telegram'a giriş yap - Basitleştirilmiş versiyon"""
    try:
        data = request.json
        action = data.get('action')
        phone = data.get('phone', '').strip()
        
        config = load_config()
        if not config['API_ID'] or not config['API_HASH']:
            return jsonify({'success': False, 'message': 'API bilgileri eksik! Lütfen önce API ID ve API Hash girin.'})
        
        session_name = 'session'
        
        async def handle_login():
            client = TelegramClient(session_name, config['API_ID'], config['API_HASH'])
            
            try:
                await client.connect()
                
                # Zaten giriş yapılmış mı kontrol et
                if await client.is_user_authorized():
                    await client.disconnect()
                    return {'success': True, 'message': 'Zaten giriş yapılmış!', 'requires_password': False}
                
                if action == 'send_code':
                    try:
                        sent_code = await client.send_code_request(phone)
                        # Session otomatik olarak phone_code_hash'i kaydeder
                        client.session.save()
                        await client.disconnect()
                        return {'success': True, 'message': 'Kod gönderildi!'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        if 'PHONE_NUMBER_INVALID' in error_msg:
                            return {'success': False, 'message': 'Telefon numarası geçersiz! Format: +90XXXXXXXXXX'}
                        elif 'FLOOD_WAIT' in error_msg:
                            return {'success': False, 'message': 'Çok fazla deneme! Lütfen birkaç dakika bekleyin.'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                
                elif action == 'verify_code':
                    code = data.get('code', '').strip()
                    if not code:
                        return {'success': False, 'message': 'Kod gerekli!'}
                    
                    try:
                        # Telethon session'dan phone_code_hash'i otomatik alır
                        result = await client.sign_in(phone, code)
                        client.session.save()
                        await client.disconnect()
                        
                        session_file = f'{session_name}.session'
                        if os.path.exists(session_file):
                            return {'success': True, 'message': 'Giriş başarılı!', 'requires_password': False}
                        else:
                            return {'success': False, 'message': 'Session kaydedilemedi. Lütfen tekrar deneyin.'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        
                        if 'PASSWORD' in error_msg or 'SESSION_PASSWORD_NEEDED' in error_msg:
                            try:
                                client.session.save()
                            except:
                                pass
                            return {'success': True, 'message': 'İki faktörlü doğrulama gerekiyor', 'requires_password': True}
                        elif 'PHONE_CODE_INVALID' in error_msg:
                            return {'success': False, 'message': 'Kod geçersiz! Lütfen doğru kodu girin.'}
                        elif 'PHONE_CODE_EXPIRED' in error_msg or 'phone_code_hash' in error_msg.lower():
                            return {'success': False, 'message': 'Kod süresi dolmuş! Lütfen yeni kod isteyin.'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                
                elif action == 'verify_password':
                    password = data.get('password', '')
                    if not password:
                        return {'success': False, 'message': 'Şifre gerekli!'}
                    
                    try:
                        await client.sign_in(password=password)
                        client.session.save()
                        await client.disconnect()
                        
                        session_file = f'{session_name}.session'
                        if os.path.exists(session_file):
                            return {'success': True, 'message': 'Giriş başarılı!'}
                        else:
                            return {'success': False, 'message': 'Session kaydedilemedi.'}
                    except Exception as e:
                        error_msg = str(e)
                        try:
                            await client.disconnect()
                        except:
                            pass
                        if 'PASSWORD' in error_msg:
                            return {'success': False, 'message': 'Şifre yanlış!'}
                        else:
                            return {'success': False, 'message': f'Hata: {error_msg}'}
                
                else:
                    return {'success': False, 'message': 'Geçersiz işlem!'}
                    
            except Exception as e:
                error_msg = str(e)
                try:
                    await client.disconnect()
                except:
                    pass
                return {'success': False, 'message': f'Hata: {error_msg}'}
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(handle_login())
        finally:
            loop.close()
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Giriş hatası: {str(e)}'})

# Bot kontrolü için global değişkenler
bot_process = None
bot_status = {'running': False, 'start_time': None, 'monitored_groups': 0, 'total_results': 0}
bot_logs = []

@app.route('/api/bot-status', methods=['GET'])
@login_required
def get_bot_status():
    """Bot durumunu getir"""
    try:
        # Session dosyası var mı kontrol et
        session_exists = os.path.exists('session.session')
        
        # Config'den grup sayısını al
        config = load_config()
        group_count = len(config.get('GROUP_IDS', []))
        
        # Bot process kontrolü
        is_running = False
        if bot_process:
            try:
                # Process hala çalışıyor mu kontrol et
                poll_result = bot_process.poll()
                if poll_result is None:
                    is_running = True
                else:
                    bot_status['running'] = False
            except:
                bot_status['running'] = False
        
        return jsonify({
            'success': True,
            'running': (bot_status['running'] and is_running) and session_exists,
            'start_time': bot_status['start_time'],
            'monitored_groups': group_count,
            'total_results': bot_status['total_results'],
            'session_exists': session_exists
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/bot-start', methods=['POST'])
@login_required
def start_bot():
    """Botu başlat"""
    import subprocess
    import threading
    from datetime import datetime
    
    global bot_process, bot_status
    
    try:
        config = load_config()
        
        # Gerekli kontroller
        if not config.get('API_ID') or not config.get('API_HASH'):
            return jsonify({'success': False, 'message': 'API bilgileri eksik! Lütfen önce API ID ve API Hash girin.'})
        
        if not config.get('GROUP_IDS') or len(config.get('GROUP_IDS', [])) == 0:
            return jsonify({'success': False, 'message': 'Grup seçilmedi! Lütfen önce gruplar sekmesinden izlemek istediğiniz grupları seçin.'})
        
        # Session kontrolü
        if not os.path.exists('session.session'):
            return jsonify({'success': False, 'message': 'Telegram girişi yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.'})
        
        if bot_status['running']:
            return jsonify({'success': False, 'message': 'Bot zaten çalışıyor!'})
        
        # Botu başlat
        bot_process = subprocess.Popen(
            ['python', 'tg_monitor.py'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            universal_newlines=True
        )
        
        bot_status['running'] = True
        bot_status['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        bot_status['monitored_groups'] = len(config.get('GROUP_IDS', []))
        bot_logs.clear()
        bot_logs.append(f"[{bot_status['start_time']}] Bot başlatıldı...")
        
        # Logları oku (thread'de)
        def read_logs():
            global bot_logs, bot_status
            try:
                for line in iter(bot_process.stdout.readline, ''):
                    if line:
                        line = line.strip()
                        bot_logs.append(line)
                        if len(bot_logs) > 1000:
                            bot_logs = bot_logs[-500:]  # Son 500 log'u tut
            except:
                pass
            finally:
                bot_status['running'] = False
                bot_logs.append(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Bot durdu.")
        
        threading.Thread(target=read_logs, daemon=True).start()
        
        return jsonify({'success': True, 'message': 'Bot başlatıldı! Logları "Bot Kontrol" sekmesinden takip edebilirsiniz.'})
        
    except Exception as e:
        bot_status['running'] = False
        return jsonify({'success': False, 'message': f'Bot başlatma hatası: {str(e)}'})

@app.route('/api/bot-stop', methods=['POST'])
@login_required
def stop_bot():
    """Botu durdur"""
    global bot_process, bot_status
    
    try:
        if not bot_status['running']:
            return jsonify({'success': False, 'message': 'Bot zaten durmuş!'})
        
        if bot_process:
            bot_process.terminate()
            try:
                bot_process.wait(timeout=5)
            except:
                bot_process.kill()
            bot_process = None
        
        bot_status['running'] = False
        bot_status['start_time'] = None
        
        return jsonify({'success': True, 'message': 'Bot durduruldu!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Bot durdurma hatası: {str(e)}'})

@app.route('/api/bot-logs', methods=['GET'])
@login_required
def get_bot_logs():
    """Bot loglarını getir"""
    try:
        return jsonify({
            'success': True,
            'logs': bot_logs[-50:]  # Son 50 log
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'logs': []})

@app.route('/api/scan', methods=['POST'])
@login_required
def start_scan():
    """Belirli tarih aralığı için tarama başlat"""
    import subprocess
    import threading
    from datetime import datetime
    
    global bot_process, bot_status
    
    try:
        data = request.json
        groups = data.get('groups', [])  # Grup listesi: [{id: ..., startDate: ..., endDate: ...}, ...]
        
        config = load_config()
        
        # Gerekli kontroller
        if not config.get('API_ID') or not config.get('API_HASH'):
            return jsonify({'success': False, 'message': 'API bilgileri eksik! Lütfen önce API ID ve API Hash girin.'})
        
        if not groups or len(groups) == 0:
            return jsonify({'success': False, 'message': 'Grup seçilmedi! Lütfen önce Ayarlar sekmesinden izlemek istediğiniz grupları seçin.'})
        
        # Session kontrolü
        if not os.path.exists('session.session'):
            return jsonify({'success': False, 'message': 'Telegram girişi yapılmamış! Lütfen önce "Gruplar" sekmesinden "Telegram\'a Giriş Yap" butonuna tıklayarak giriş yapın.'})
        
        if bot_status['running']:
            return jsonify({'success': False, 'message': 'Bot zaten çalışıyor! Önce botu durdurun.'})
        
        # Grup ID'lerini ve tarih aralıklarını config'e kaydet
        group_ids = []
        for group in groups:
            group_id = group.get('id')
            if group_id:
                group_ids.append({
                    'id': group_id,
                    'startDate': group.get('startDate'),
                    'endDate': group.get('endDate')
                })
        
        if not group_ids:
            return jsonify({'success': False, 'message': 'Geçerli grup bulunamadı!'})
        
        # Config'i güncelle
        config['GROUP_IDS'] = group_ids
        save_config(config)
        
        # Botu başlat
        bot_process = subprocess.Popen(
            ['python', 'tg_monitor.py'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            universal_newlines=True
        )
        
        bot_status['running'] = True
        bot_status['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        bot_status['monitored_groups'] = len(group_ids)
        bot_logs.clear()
        bot_logs.append(f"[{bot_status['start_time']}] Tarama başlatıldı...")
        
        # Grup bilgilerini log'a ekle
        for group in group_ids:
            date_info = ""
            if group.get('startDate') or group.get('endDate'):
                date_info = f" (Tarih: {group.get('startDate', 'Başlangıç yok')} - {group.get('endDate', 'Bitiş yok')})"
            bot_logs.append(f"Grup {group.get('id')}{date_info}")
        
        # Logları oku (thread'de)
        def read_logs():
            global bot_logs, bot_status
            try:
                for line in iter(bot_process.stdout.readline, ''):
                    if line:
                        line = line.strip()
                        bot_logs.append(line)
                        if len(bot_logs) > 1000:
                            bot_logs = bot_logs[-500:]  # Son 500 log'u tut
            except:
                pass
            finally:
                bot_status['running'] = False
                bot_logs.append(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Tarama tamamlandı.")
        
        threading.Thread(target=read_logs, daemon=True).start()
        
        return jsonify({'success': True, 'message': f'Tarama başlatıldı! ({len(group_ids)} grup)'})
        
    except Exception as e:
        bot_status['running'] = False
        return jsonify({'success': False, 'message': f'Tarama başlatma hatası: {str(e)}'})

@app.route('/api/scan-status', methods=['GET'])
@login_required
def get_scan_status():
    """Tarama durumunu getir"""
    try:
        global bot_status, bot_process, bot_logs
        
        is_running = False
        if bot_process:
            try:
                poll_result = bot_process.poll()
                if poll_result is None:
                    is_running = True
                else:
                    bot_status['running'] = False
            except:
                bot_status['running'] = False
        
        # Son 50 log'u gönder
        recent_logs = bot_logs[-50:] if bot_logs else []
        
        # Results dosyasından sonuç sayısını hesapla
        result_count = 0
        try:
            if os.path.exists('results.txt'):
                with open('results.txt', 'r', encoding='utf-8') as f:
                    content = f.read()
                    # Her "=====" ile başlayan blok bir sonuçtur
                    result_count = content.count('=' * 60)
        except:
            pass
        
        # Son 50 log'u gönder
        recent_logs = bot_logs[-50:] if bot_logs else []
        
        # Sonuç sayısını kontrol et
        config = load_config()
        results_file = config.get('RESULTS_FILE', 'results.txt')
        result_count = 0
        if os.path.exists(results_file):
            with open(results_file, 'r', encoding='utf-8') as f:
                content = f.read()
                # Sonuç sayısını hesapla (her sonuç "=" ile başlar)
                result_count = content.count('=' * 60)
        
        return jsonify({
            'success': True,
            'running': bot_status['running'] and is_running,
            'start_time': bot_status['start_time'],
            'monitored_groups': bot_status.get('monitored_groups', 0),
            'result_count': result_count,
            'logs': recent_logs,
            'log_count': len(bot_logs)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    print("🌐 Web paneli başlatılıyor...")
    print(f"📱 Tarayıcıda http://localhost:{port} adresine gidin")
    app.run(debug=False, host='0.0.0.0', port=port)

